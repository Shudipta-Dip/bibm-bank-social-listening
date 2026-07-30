"""Reorder your_label after text; write Excel with dropdown."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "out"
df = pd.read_csv(OUT / "edge_cases_for_human.csv")

cols = [c for c in df.columns if c != "your_label"]
ti = cols.index("text")
cols.insert(ti + 1, "your_label")
df = df.reindex(columns=cols)
df["your_label"] = df["your_label"].fillna("")

df.to_csv(OUT / "edge_cases_for_human.csv", index=False, encoding="utf-8-sig")

wb = Workbook()
ws = wb.active
ws.title = "edge_cases"
for row in dataframe_to_rows(df, index=False, header=True):
    ws.append(list(row))

header = [c.value for c in ws[1]]
yl_col = header.index("your_label") + 1
yl_letter = get_column_letter(yl_col)
n = max(ws.max_row, 2)

dv = DataValidation(
    type="list",
    formula1='"positive,negative,neutral"',
    allow_blank=True,
    showDropDown=False,  # Excel: False = show the arrow
)
dv.error = "Pick positive, negative, or neutral"
dv.errorTitle = "Invalid label"
dv.prompt = "Select sentiment"
dv.promptTitle = "your_label"
ws.add_data_validation(dv)
dv.add(f"{yl_letter}2:{yl_letter}{n}")

text_col = header.index("text") + 1
ws.column_dimensions[get_column_letter(text_col)].width = 60
ws.column_dimensions[yl_letter].width = 14
for i, name in enumerate(header, 1):
    if name == "why_suspect":
        ws.column_dimensions[get_column_letter(i)].width = 40
    elif name not in ("text", "your_label"):
        ws.column_dimensions[get_column_letter(i)].width = min(18, max(10, len(str(name)) + 2))

xlsx = OUT / "edge_cases_for_human.xlsx"
wb.save(xlsx)
print("columns:", list(df.columns))
print("wrote", xlsx)
